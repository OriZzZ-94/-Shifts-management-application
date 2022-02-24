import openpyxl
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from datetime import datetime
from Classes import Employee


# This function purpose is to import the data of the employees and return the list of Employee objects
def importList():
    employee_list = []
    Employees = openpyxl.load_workbook("Employees.xlsx")
    sheet = Employees.worksheets[0]

    number_of_employees = sheet.cell(row=2, column=1).value
    for x in range(number_of_employees):
        name = sheet.cell(row=(x + 1) * 5, column=1).value
        employee_number = int(sheet.cell(row=(x + 1) * 5, column=2).value)
        num_of_shifts = int(sheet.cell(row=(x + 1) * 5 + 1, column=2).value)

        employee_list.append(Employee(name, num_of_shifts, employee_number))

    Employees.close()
    return employee_list


# This function deletes an employee
def delete_employee(employee_number):
    Employees = openpyxl.load_workbook("Employees.xlsx")
    sheet = Employees.worksheets[0]
    number_of_employees = sheet.cell(row=2, column=1).value
    sheet['A2'] = sheet.cell(row=2, column=1).value - 1
    for i in range(employee_number + 1, number_of_employees + 1):
        sheet.cell(row=i * 5, column=2).value = sheet.cell(row=i * 5, column=2).value - 1
    sheet.delete_rows(employee_number * 5, 5)
    Employees.save("Employees.xlsx")
    Employees.close()


def edit_employee(employee1):
    Employees = openpyxl.load_workbook("Employees.xlsx")
    sheet = Employees.worksheets[0]
    i = int(employee1.employee_number)
    sheet.cell(row=i * 5 + 1, column=2).value = employee1.num_of_shifts
    Employees.save("Employees.xlsx")
    Employees.close()


def add_employee(name, week_shifts):
    Employees = openpyxl.load_workbook("Employees.xlsx")
    sheet = Employees.worksheets[0]
    number_of_employees = sheet.cell(row=2, column=1).value
    i = int(number_of_employees) + 1
    sheet.cell(row=i * 5, column=1).value = name
    sheet.cell(row=i * 5, column=2).value = i
    sheet.cell(row=i * 5 + 1, column=1).value = "Number of shifts"
    sheet.cell(row=i * 5 + 1, column=2).value = week_shifts
    sheet.cell(row=2, column=1).value = sheet.cell(row=2, column=1).value + 1
    Employees.save("Employees.xlsx")
    Employees.close()


def save_export(weekly_rest, emp_list):
    Employees = openpyxl.load_workbook("Employees.xlsx")

    # Creating and printing the new scheudle
    date = datetime.today()
    date = str(date).split(" ")[0]
    Employees.create_sheet(date)
    new_schedule = Employees.get_sheet_by_name(date)
    new_schedule['A3'] = "בוקר"
    new_schedule['B3'] = "אחמש"
    new_schedule['B4'] = "עובד"
    new_schedule['A5'] = "צהריים"
    new_schedule['B5'] = "אחמש"
    new_schedule['B6'] = "עובד"
    new_schedule['A7'] = "לילה"
    new_schedule['B7'] = "אחמש"
    new_schedule['B8'] = "עובד"
    new_schedule['C2'] = "ראשון"
    new_schedule['D2'] = "שני"
    new_schedule['E2'] = "שלישי"
    new_schedule['F2'] = "רביעי"
    new_schedule['G2'] = "חמישי"
    new_schedule['H2'] = "שישי"
    new_schedule['I2'] = "שבת"

    column = 3
    row = 3
    for x in range(42):
        if weekly_rest[x] == 0:
            new_schedule.cell(row=row, column=column).value = ""
        else:
            new_schedule.cell(row=row, column=column).value = emp_list[weekly_rest[x] - 1].name
        row += 1
        if row % 9 == 0:
            row = 3
            column += 1

    thin_border = Border(
        left=Side(border_style=BORDER_THIN, color='00000000'),
        right=Side(border_style=BORDER_THIN, color='00000000'),
        top=Side(border_style=BORDER_THIN, color='00000000'),
        bottom=Side(border_style=BORDER_THIN, color='00000000'))
    for x in range(1, 10):
        for y in range(1, 9):
            new_schedule.cell(row=y, column=x).border = thin_border

    # Updating the employees stats
    employees = Employees.worksheets[0]
    number_of_employees = employees.cell(row=2, column=1).value
    updated_list = []
    for x in range(number_of_employees):
        updated_list.append(0)
    for x in range(42):
        if (x % 6 == 4 or x % 6 == 5) and not weekly_rest[x] == 0:
            updated_list[weekly_rest[x] - 1] += 1

    for x in range(number_of_employees):
        employees.cell(row=(x + 1) * 5 + 2, column=2).value = updated_list[x]
        employees.cell(row=(x + 1) * 5 + 3, column=4).value = employees.cell(row=(x + 1) * 5 + 3, column=2).value

    Employees.save("Employees.xlsx")
    Employees.close()
