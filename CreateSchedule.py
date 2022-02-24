import openpyxl
from tkinter import *
from tkinter.ttk import *
from Edit_Emp import values
import EmployeeData


def edit_sc(edit_sc_win):
    # saving changes
    def save_schedule():
        Employees = openpyxl.load_workbook("Employees.xlsx")
        schedule = Employees.worksheets[1]
        updated_list = []
        for x in range(len(emp_list)):
            updated_list.append(0)

        row1 = 3
        column1 = 3
        for x in range(42):
            if shifts[x].get().split(':')[0] == '0':
                schedule.cell(row=row1, column=column1).value = ""
            else:
                schedule.cell(row=row1, column=column1).value = shifts[x].get().split(':')[1][1:]
            if (row1 == 7 or row1 == 8) and not shifts[x].get() == "":
                updated_list[int(shifts[x].get().split(':')[0]) - 1] += 1
            row1 += 1
            if row1 % 9 == 0:
                row1 = 3
                column1 += 1

        employees = Employees.worksheets[0]

        Employees.save("Employees.xlsx")
        Employees.close()

    # Loading the Latest schedule
    emp_list = EmployeeData.importList()
    headline = Label(edit_sc_win, font="Ariel 14 bold underline", anchor='center', text="Edit Active Schedule")
    headline.grid(row=0, pady=10)

    # main frame for hand picking shifts
    frame_body = Frame(edit_sc_win, relief='groove', width=500, height=500)
    frame_body.grid(row=1, padx=10, pady=10)

    # frame for the days of the week on top
    frame_body_head = Frame(frame_body, relief='groove')
    frame_body_head.grid(row=0, column=1)
    lbl_sun = Label(frame_body_head, text="Sunday").grid(row=0, column=1, padx=10, pady=10)
    lbl_mon = Label(frame_body_head, text="Monday").grid(row=0, column=2, padx=10, pady=10)
    lbl_teu = Label(frame_body_head, text="Tuesday").grid(row=0, column=3, padx=10, pady=10)
    lbl_wed = Label(frame_body_head, text="Wednesday").grid(row=0, column=4, padx=10, pady=10)
    lbl_thu = Label(frame_body_head, text="Thursday").grid(row=0, column=5, padx=10, pady=10)
    lbl_fri = Label(frame_body_head, text="Friday").grid(row=0, column=6, padx=10, pady=10)
    lbl_sat = Label(frame_body_head, text="Saturday").grid(row=0, column=7, padx=10, pady=10)

    # frame for days and positions
    frame_body_left = Frame(frame_body, relief='groove')
    frame_body_left.grid(row=2, column=0, sticky='w')
    lbl_mor = Label(frame_body_left, text="Morning:").grid(row=0, column=0, padx=10, pady=10)
    lbl_after = Label(frame_body_left, text="Afternoon:").grid(row=2, column=0, padx=10, pady=10)
    lbl_eve = Label(frame_body_left, text="Night:").grid(row=4, column=0, padx=10, pady=10)
    for x in range(0, 5, 2):
        lbl_incharge = Label(frame_body_left, text="Incharge")
        lbl_worker = Label(frame_body_left, text="Worker")
        lbl_incharge.grid(row=x, column=1, padx=2, pady=2)
        lbl_worker.grid(row=x + 1, column=1, padx=2, pady=2)


    frame_body_main = Frame(frame_body, relief='groove')
    frame_body_main.grid(row=2, column=1)
    shifts = []
    row1 = 0
    column1 = 0
    for x in range(0, 42, 2):
        combo_shift = Combobox(frame_body_main, state='readonly', width=7)
        combo_shift.configure(value=values(emp_list))
        shifts.append(combo_shift)
        shifts[x].grid(row=row1, column=column1, padx=2, pady=5)
        combo_shift2 = Combobox(frame_body_main, state='readonly', width=7)
        combo_shift2.configure(value=values(emp_list))
        shifts.append(combo_shift2)
        shifts[x + 1].grid(row=row1 + 1, column=column1, padx=2, pady=5)
        shifts[x].current(0)
        shifts[x + 1].current(0)
        row1 += 2
        if row1 % 6 == 0:
            row1 = 0
            column1 += 1


    Employees = openpyxl.load_workbook("Employees.xlsx")
    schedule = Employees.worksheets[1]
    column1 = 3
    row1 = 3
    for x in range(42):
        for y in emp_list:
            if schedule.cell(row=row1, column=column1).value == y.name:
                shifts[x].current(y.employee_number)
        row1 += 1
        if row1 % 9 == 0:
            row1 = 3
            column1 += 1
    Employees.close()

    btn_save = Button(edit_sc_win, text="Save changes", command=save_schedule)
    btn_save.grid(row=3, column=0, pady=10)
